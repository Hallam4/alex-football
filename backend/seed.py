#!/usr/bin/env python3
"""Seed script: Excel → SQLite for Alex Football app."""

import os
import re
import sys
from datetime import datetime
from collections import defaultdict

import openpyxl
from sqlalchemy import create_engine, func
from sqlalchemy.orm import Session

sys.path.insert(0, os.path.dirname(__file__))
from db_models import (
    Base, Player, Block, BlockRoster, GameResult,
    HeadToHead, MomAward, LeagueStanding,
)

DB_PATH = os.path.join(os.path.dirname(__file__), "data", "football.db")
HALLAM_FILE = os.path.expanduser("~/Desktop/Block Football_Hallam.xlsx")
STATS_FILE = os.path.expanduser("~/Desktop/Football Stats_2026_new.xlsm")


def parse_result_cell(cell_value):
    """Parse W/L/D result cell, handling sub notation like 'W - SubName'.
    Returns (result, is_sub, sub_name) or (None, False, None)."""
    if not cell_value:
        return None, False, None
    val = str(cell_value).strip()
    if val in ("W", "L", "D"):
        return val, False, None
    m = re.match(r"^([WLD])\s*[-–]\s*(.+)$", val)
    if m:
        sub_name = m.group(2).strip()
        if sub_name.lower() == "guest":
            return m.group(1), False, None
        return m.group(1), True, sub_name
    return None, False, None


def parse_six_a_side(ws):
    """Parse '6 a side' sheet into structured block data."""
    block_headers = []
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=2).value
        if val and isinstance(val, str) and val.strip().startswith("Block"):
            m = re.search(r"Block\s+(\d+)", val.strip())
            if m:
                block_headers.append({"row": r, "number": int(m.group(1)), "header": val.strip()})

    blocks = []
    for idx, bh in enumerate(block_headers):
        header_row = bh["row"]
        end_row = block_headers[idx + 1]["row"] if idx + 1 < len(block_headers) else ws.max_row + 1

        # Names row is consistently header_row + 3
        names_row = header_row + 3
        player_columns = {}
        for col in range(3, 15):  # C-N
            name = ws.cell(row=names_row, column=col).value
            if name and isinstance(name, str) and name.strip():
                player_columns[col] = name.strip()

        if not player_columns:
            print(f"  Warning: No players for Block {bh['number']} (row {names_row})")
            continue

        matches = []
        week_counter = 0

        for r in range(names_row + 1, end_row):
            date_val = ws.cell(row=r, column=2).value
            if not isinstance(date_val, datetime):
                continue

            results = []
            valid_count = 0
            for col, regular_name in player_columns.items():
                result, is_sub, sub_name = parse_result_cell(ws.cell(row=r, column=col).value)
                if result:
                    valid_count += 1
                    results.append({
                        "regular_name": regular_name,
                        "result": result,
                        "is_sub": is_sub,
                        "sub_name": sub_name,
                        "player_name": sub_name if is_sub else regular_name,
                    })

            if valid_count < 6:
                continue

            week_counter += 1
            wn = ws.cell(row=r, column=1).value
            week_num = int(wn) if isinstance(wn, (int, float)) else week_counter

            score_w = ws.cell(row=r, column=26).value  # Z
            score_l = ws.cell(row=r, column=27).value  # AA

            matches.append({
                "date": date_val.date(),
                "week_number": week_num,
                "results": results,
                "score_w": int(score_w) if isinstance(score_w, (int, float)) else None,
                "score_l": int(score_l) if isinstance(score_l, (int, float)) else None,
            })

        blocks.append({
            "number": bh["number"],
            "header": bh["header"],
            "start_date": matches[0]["date"] if matches else None,
            "player_columns": player_columns,
            "matches": matches,
        })
        print(f"  Block {bh['number']}: {len(matches)} matches, {len(player_columns)} regulars")

    return blocks


def parse_to_share(ws):
    """Parse 'To Share' sheet for current league standings."""
    standings = []
    for r in range(5, ws.max_row + 1):
        pos = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        if not pos or not name or not isinstance(pos, (int, float)):
            continue
        standings.append({
            "position": int(pos),
            "name": str(name).strip(),
            "played": int(ws.cell(row=r, column=3).value or 0),
            "won": int(ws.cell(row=r, column=4).value or 0),
            "drawn": int(ws.cell(row=r, column=5).value or 0),
            "lost": int(ws.cell(row=r, column=6).value or 0),
            "mom": int(ws.cell(row=r, column=7).value or 0),
            "points": float(ws.cell(row=r, column=8).value or 0),
            "gd": int(ws.cell(row=r, column=9).value or 0),
            "ppg": round(float(ws.cell(row=r, column=10).value or 0), 3),
        })
    return standings


def parse_h2h(ws):
    """Parse 'Player on Player' columns AB-AI for H2H (same-team synergy) records."""
    records = []
    for r in range(2, ws.max_row + 1):
        p1 = ws.cell(row=r, column=29).value  # AC
        p2 = ws.cell(row=r, column=30).value  # AD
        if not p1 or not p2:
            continue
        played = ws.cell(row=r, column=31).value
        if not played or not isinstance(played, (int, float)):
            continue
        records.append({
            "player_a": str(p1).strip(),
            "player_b": str(p2).strip(),
            "played": int(played),
            "wins": int(ws.cell(row=r, column=32).value or 0),
            "losses": int(ws.cell(row=r, column=33).value or 0),
            "draws": int(ws.cell(row=r, column=34).value or 0),
            "goals_scored": int(ws.cell(row=r, column=35).value or 0),
        })
    return records


def parse_mom(ws):
    """Parse 'MoM' sheet for Man of the Match awards."""
    awards = []
    for r in range(2, ws.max_row + 1):
        week = ws.cell(row=r, column=2).value
        name = ws.cell(row=r, column=5).value
        if not name or not week:
            continue
        date_val = ws.cell(row=r, column=3).value
        awards.append({
            "week_number": int(week),
            "game_date": date_val.date() if isinstance(date_val, datetime) else None,
            "player_name": str(name).strip(),
            "votes": int(ws.cell(row=r, column=6).value or 0),
            "score": float(ws.cell(row=r, column=7).value or 0),
        })
    return awards


def seed():
    print("Loading Excel files...")
    wb_hallam = openpyxl.load_workbook(HALLAM_FILE, data_only=True)
    wb_stats = openpyxl.load_workbook(STATS_FILE, data_only=True)

    print("\nParsing '6 a side' sheet...")
    blocks_data = parse_six_a_side(wb_hallam["6 a side"])

    print(f"\nParsing 'To Share' sheet...")
    league_data = parse_to_share(wb_stats["To Share"])
    print(f"  {len(league_data)} players")

    print("\nParsing 'Player on Player' for H2H...")
    h2h_data = parse_h2h(wb_stats["Player on Player"])
    print(f"  {len(h2h_data)} records")

    print("\nParsing 'MoM' sheet...")
    mom_data = parse_mom(wb_stats["MoM"])
    print(f"  {len(mom_data)} awards")

    # --- Create database ---
    print(f"\nCreating database at {DB_PATH}...")
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)

    engine = create_engine(f"sqlite:///{DB_PATH}", echo=False)
    Base.metadata.create_all(engine)

    with Session(engine) as session:
        players_cache: dict[str, int] = {}

        def get_or_create_player(name: str) -> int:
            clean = name.strip()
            if clean in players_cache:
                return players_cache[clean]
            p = Player(name=clean, is_active=False)
            session.add(p)
            session.flush()
            players_cache[clean] = p.id
            return p.id

        latest_block = max(b["number"] for b in blocks_data) if blocks_data else 0

        # Determine active players
        active_names: set[str] = set()
        for b in blocks_data:
            if b["number"] == latest_block:
                active_names.update(b["player_columns"].values())
                for m in b["matches"]:
                    for r in m["results"]:
                        if r["is_sub"] and r["sub_name"]:
                            active_names.add(r["sub_name"])
        for entry in league_data:
            active_names.add(entry["name"])

        # Import blocks and game results
        print("\nImporting blocks and game results...")
        total_games = 0

        for bd in blocks_data:
            block = Block(
                id=bd["number"],
                name=f"Block {bd['number']}",
                start_date=bd["start_date"],
                quarter=((bd["number"] - 2) % 4) + 1,
            )
            session.add(block)
            session.flush()

            for name in bd["player_columns"].values():
                pid = get_or_create_player(name)
                session.add(BlockRoster(block_id=block.id, player_id=pid, is_regular=True))

            fill_ins: set[str] = set()

            for match in bd["matches"]:
                for r in match["results"]:
                    pid = get_or_create_player(r["player_name"])
                    replaced_pid = None
                    if r["is_sub"]:
                        replaced_pid = get_or_create_player(r["regular_name"])
                        fill_ins.add(r["player_name"])

                    gf, ga = None, None
                    if match["score_w"] is not None and match["score_l"] is not None:
                        if r["result"] == "W":
                            gf, ga = match["score_w"], match["score_l"]
                        elif r["result"] == "L":
                            gf, ga = match["score_l"], match["score_w"]
                        elif r["result"] == "D":
                            gf = ga = match["score_w"]

                    session.add(GameResult(
                        block_id=block.id,
                        week_number=match["week_number"],
                        game_date=match["date"],
                        player_id=pid,
                        result=r["result"],
                        is_sub=r["is_sub"],
                        replaced_player_id=replaced_pid,
                        goals_for=gf,
                        goals_against=ga,
                    ))
                    total_games += 1

            for name in fill_ins:
                pid = players_cache[name]
                exists = session.query(BlockRoster).filter_by(
                    block_id=block.id, player_id=pid
                ).first()
                if not exists:
                    session.add(BlockRoster(block_id=block.id, player_id=pid, is_regular=False))

        print(f"  {total_games} game results")

        # Player metadata
        print("Setting player metadata...")
        for player in session.query(Player).all():
            player.is_active = player.name in active_names
            earliest = (
                session.query(GameResult.game_date)
                .filter(GameResult.player_id == player.id, GameResult.game_date.isnot(None))
                .order_by(GameResult.game_date)
                .first()
            )
            if earliest:
                player.first_game_date = earliest[0]

        # League standings
        print("Importing league standings...")
        for entry in league_data:
            pid = get_or_create_player(entry["name"])
            session.add(LeagueStanding(
                block_id=latest_block,
                player_id=pid,
                played=entry["played"],
                won=entry["won"],
                drawn=entry["drawn"],
                lost=entry["lost"],
                points=entry["points"],
                goal_difference=entry["gd"],
                mom_bonus=entry["mom"],
                ppg=entry["ppg"],
            ))

        # H2H
        print("Importing H2H records...")
        h2h_count = 0
        for rec in h2h_data:
            pid_a = get_or_create_player(rec["player_a"])
            pid_b = get_or_create_player(rec["player_b"])
            session.add(HeadToHead(
                player_a_id=pid_a, player_b_id=pid_b,
                played=rec["played"], wins=rec["wins"],
                draws=rec["draws"], losses=rec["losses"],
                goals_scored=rec["goals_scored"],
            ))
            h2h_count += 1
        print(f"  {h2h_count} H2H records")

        # MoM
        print("Importing MoM awards...")
        for award in mom_data:
            pid = get_or_create_player(award["player_name"])
            session.add(MomAward(
                block_id=latest_block,
                game_date=award["game_date"],
                week_number=award["week_number"],
                player_id=pid,
                votes=award["votes"],
                score=award["score"],
            ))

        session.commit()

        # Summary
        player_count = session.query(Player).count()
        block_count = session.query(Block).count()
        game_count = session.query(GameResult).count()
        active_count = session.query(Player).filter(Player.is_active == True).count()

        print(f"\n{'=' * 40}")
        print("Seed complete!")
        print(f"  Players: {player_count} ({active_count} active)")
        print(f"  Blocks: {block_count}")
        print(f"  Game results: {game_count}")
        print(f"  League standings: {len(league_data)}")
        print(f"  H2H records: {h2h_count}")
        print(f"  MoM awards: {len(mom_data)}")
        print(f"{'=' * 40}")

        # Verification
        print("\nVerification:")
        hallam = session.query(Player).filter(Player.name == "Hallam").first()
        if hallam:
            total = session.query(func.count(GameResult.id)).filter(
                GameResult.player_id == hallam.id
            ).scalar()
            wins = session.query(func.count(GameResult.id)).filter(
                GameResult.player_id == hallam.id, GameResult.result == "W"
            ).scalar()
            wr = (wins / total * 100) if total > 0 else 0
            print(f"  Hallam: {total} games, {wins} wins, {wr:.1f}%")
            print(f"  Expected: ~123 games, ~47 wins, ~38.2%")


if __name__ == "__main__":
    seed()
